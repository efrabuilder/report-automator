# app.py — Report Automator Web Interface
# Efraín Rojas Artavia

import io
import os
import base64
import tempfile
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
from collections import defaultdict

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)

PALETTE = ["#FF6B35","#7C3AED","#10B981","#3B82F6","#F59E0B",
           "#EF4444","#8B5CF6","#06B6D4","#84CC16","#F97316"]

SAMPLE_CSV = """Month,Region,Category,Revenue,Expenses,Units
Jan,North,Electronics,45200,31000,312
Jan,South,Clothing,28400,18500,540
Jan,East,Electronics,38900,26700,275
Feb,North,Electronics,51200,34200,358
Feb,South,Clothing,31800,20100,602
Feb,East,Food,22100,14300,1020
Mar,North,Clothing,33400,21800,634
Mar,South,Electronics,48700,33100,342
Mar,East,Electronics,43200,29600,305
Apr,North,Food,21500,13900,980
Apr,South,Electronics,55100,37400,387
Apr,East,Clothing,29800,19400,566
May,North,Electronics,58400,39700,410
May,South,Food,24200,15700,1100
May,East,Electronics,50100,34300,352
Jun,North,Clothing,36700,23900,696
Jun,South,Electronics,61200,41600,430
Jun,East,Food,26800,17400,1220"""


def make_chart(df, chart_cfg):
    fig, ax = plt.subplots(figsize=(8, 4), facecolor="#1a1a1a")
    ax.set_facecolor("#1a1a1a")
    kind  = chart_cfg.get("type", "bar")
    title = chart_cfg.get("title", "Chart")

    if kind == "bar":
        x_col, y_col = chart_cfg.get("x"), chart_cfg.get("y")
        grouped = df.groupby(x_col)[y_col].sum().reset_index()
        bars = ax.bar(grouped[x_col].astype(str), grouped[y_col],
                      color=PALETTE[:len(grouped)], edgecolor="none", width=0.6)
        ax.bar_label(bars, fmt="%.0f", padding=4, color="#F0EEE8", fontsize=8)

    elif kind == "line":
        x_col  = chart_cfg.get("x")
        y_cols = chart_cfg.get("y_cols", [])
        grouped = df.groupby(x_col)[y_cols].sum().reset_index()
        for i, col in enumerate(y_cols):
            ax.plot(grouped[x_col].astype(str), grouped[col],
                    marker="o", color=PALETTE[i], linewidth=2, markersize=5, label=col)
        ax.legend(facecolor="#2a2a2a", edgecolor="none", labelcolor="#F0EEE8", fontsize=8)
        plt.xticks(rotation=30, ha="right")

    elif kind == "pie":
        label_col = chart_cfg.get("label")
        value_col = chart_cfg.get("value")
        grouped = df.groupby(label_col)[value_col].sum().reset_index()
        wedges, texts, autotexts = ax.pie(
            grouped[value_col], labels=grouped[label_col].astype(str),
            autopct="%1.1f%%", colors=PALETTE[:len(grouped)],
            startangle=140, wedgeprops={"edgecolor":"#1a1a1a","linewidth":2}
        )
        for t in texts: t.set_color("#888")
        for a in autotexts: a.set_color("#F0EEE8"); a.set_fontsize(8)

    ax.set_title(title, color="#F0EEE8", fontsize=12, fontweight="bold", pad=12)
    ax.tick_params(colors="#888")
    ax.spines[:].set_visible(False)
    if kind != "pie":
        ax.yaxis.grid(True, color="#2a2a2a", linewidth=0.8)
        ax.set_axisbelow(True)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight", facecolor="#1a1a1a")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


def build_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False
    hfill = PatternFill("solid", fgColor="FF6B35")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    afill = PatternFill("solid", fgColor="1E1E1E")

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, col)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col))+4, 14)

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font = Font(color="F0EEE8", size=10)
            if ri % 2 == 0: c.fill = afill
            c.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/sample")
def sample():
    return jsonify({"csv": SAMPLE_CSV})


@app.route("/api/analyze", methods=["POST"])
def analyze():
    data = request.get_json()
    csv_text = data.get("csv", "").strip()
    charts_cfg = data.get("charts", [])

    if not csv_text:
        return jsonify({"error": "No CSV data provided"}), 400

    try:
        df = pd.read_csv(io.StringIO(csv_text))
    except Exception as e:
        return jsonify({"error": f"Could not parse CSV: {e}"}), 400

    # Stats
    numeric = df.select_dtypes(include="number")
    stats = {}
    for col in numeric.columns:
        stats[col] = {
            "sum":  round(numeric[col].sum(), 2),
            "mean": round(numeric[col].mean(), 2),
            "max":  round(numeric[col].max(), 2),
            "min":  round(numeric[col].min(), 2),
        }

    # Charts
    charts = []
    default_charts = [
        {"type":"bar",  "title":"Revenue by Region", "x":"Region", "y":"Revenue"},
        {"type":"line", "title":"Monthly Trend",      "x":"Month",  "y_cols":["Revenue","Expenses"]},
        {"type":"pie",  "title":"Revenue Distribution","label":"Category","value":"Revenue"},
    ] if not charts_cfg else charts_cfg

    for cfg in default_charts:
        try:
            img = make_chart(df, cfg)
            charts.append({"title": cfg.get("title","Chart"), "image": img})
        except Exception as e:
            charts.append({"title": cfg.get("title","Chart"), "error": str(e)})

    return jsonify({
        "rows":    len(df),
        "columns": list(df.columns),
        "stats":   stats,
        "charts":  charts,
        "preview": df.head(10).to_dict(orient="records"),
    })


@app.route("/api/export", methods=["POST"])
def export():
    data = request.get_json()
    csv_text = data.get("csv", "").strip()
    if not csv_text:
        return jsonify({"error": "No CSV data"}), 400
    try:
        df = pd.read_csv(io.StringIO(csv_text))
        buf = build_excel(df)
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
