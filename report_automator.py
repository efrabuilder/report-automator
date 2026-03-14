"""
Report Automator — Efraín Rojas Artavia
Generates Excel + PDF reports with charts from CSV/Excel data,
and sends them via email on a configurable schedule.
"""

import os
import smtplib
import schedule
import time
import logging
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from fpdf import FPDF

from config import CONFIG

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("report_automator.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)
CHART_DIR = OUTPUT_DIR / "charts"
CHART_DIR.mkdir(exist_ok=True)


# ── Data Loading ───────────────────────────────────────────────────────────────
def load_data(path: str) -> pd.DataFrame:
    """Load CSV or Excel file into a DataFrame."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Data file not found: {path}")
    if p.suffix.lower() == ".csv":
        df = pd.read_csv(p)
    elif p.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(p)
    else:
        raise ValueError(f"Unsupported file format: {p.suffix}")
    log.info(f"Loaded {len(df)} rows from {p.name}")
    return df


# ── Chart Generation ───────────────────────────────────────────────────────────
PALETTE = ["#FF6B35", "#7C3AED", "#10B981", "#3B82F6", "#F59E0B",
           "#EF4444", "#8B5CF6", "#06B6D4", "#84CC16", "#F97316"]

def _save_fig(fig, name: str) -> Path:
    path = CHART_DIR / f"{name}.png"
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="#1a1a1a")
    plt.close(fig)
    log.info(f"Chart saved: {path.name}")
    return path


def chart_bar(df: pd.DataFrame, x_col: str, y_col: str, title: str, name: str) -> Path:
    fig, ax = plt.subplots(figsize=(10, 5), facecolor="#1a1a1a")
    ax.set_facecolor("#1a1a1a")
    bars = ax.bar(df[x_col].astype(str), df[y_col], color=PALETTE[:len(df)], edgecolor="none", width=0.6)
    ax.bar_label(bars, fmt="%.1f", padding=4, color="#F0EEE8", fontsize=9)
    ax.set_title(title, color="#F0EEE8", fontsize=13, fontweight="bold", pad=14)
    ax.set_xlabel(x_col, color="#888", fontsize=10)
    ax.set_ylabel(y_col, color="#888", fontsize=10)
    ax.tick_params(colors="#888")
    ax.spines[:].set_visible(False)
    ax.yaxis.grid(True, color="#2a2a2a", linewidth=0.8)
    ax.set_axisbelow(True)
    return _save_fig(fig, name)


def chart_line(df: pd.DataFrame, x_col: str, y_cols: list, title: str, name: str) -> Path:
    fig, ax = plt.subplots(figsize=(10, 5), facecolor="#1a1a1a")
    ax.set_facecolor("#1a1a1a")
    for i, col in enumerate(y_cols):
        ax.plot(df[x_col].astype(str), df[col], marker="o", color=PALETTE[i],
                linewidth=2, markersize=5, label=col)
    ax.set_title(title, color="#F0EEE8", fontsize=13, fontweight="bold", pad=14)
    ax.tick_params(colors="#888")
    ax.spines[:].set_visible(False)
    ax.yaxis.grid(True, color="#2a2a2a", linewidth=0.8)
    ax.set_axisbelow(True)
    ax.legend(facecolor="#2a2a2a", edgecolor="none", labelcolor="#F0EEE8", fontsize=9)
    plt.xticks(rotation=30, ha="right")
    return _save_fig(fig, name)


def chart_pie(df: pd.DataFrame, label_col: str, value_col: str, title: str, name: str) -> Path:
    fig, ax = plt.subplots(figsize=(7, 7), facecolor="#1a1a1a")
    wedges, texts, autotexts = ax.pie(
        df[value_col], labels=df[label_col].astype(str),
        autopct="%1.1f%%", colors=PALETTE[:len(df)],
        pctdistance=0.82, startangle=140,
        wedgeprops={"edgecolor": "#1a1a1a", "linewidth": 2}
    )
    for t in texts: t.set_color("#888")
    for a in autotexts: a.set_color("#F0EEE8"); a.set_fontsize(9)
    ax.set_title(title, color="#F0EEE8", fontsize=13, fontweight="bold", pad=14)
    return _save_fig(fig, name)


def generate_charts(df: pd.DataFrame) -> list:
    """Generate all charts defined in CONFIG and return their paths."""
    paths = []
    for i, ch in enumerate(CONFIG.get("charts", [])):
        kind = ch.get("type", "bar")
        title = ch.get("title", f"Chart {i+1}")
        name = f"chart_{i+1}"
        try:
            if kind == "bar":
                p = chart_bar(df, ch["x"], ch["y"], title, name)
            elif kind == "line":
                p = chart_line(df, ch["x"], ch["y_cols"], title, name)
            elif kind == "pie":
                p = chart_pie(df, ch["label"], ch["value"], title, name)
            else:
                log.warning(f"Unknown chart type '{kind}', skipping.")
                continue
            paths.append(p)
        except Exception as e:
            log.error(f"Chart '{title}' failed: {e}")
    return paths


# ── Excel Report ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="FF6B35")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="1E1E1E")
TITLE_FONT  = Font(bold=True, size=14, color="FF6B35")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="333333"),
    top=Side(style="thin", color="333333"),
)

def build_excel(df: pd.DataFrame, chart_paths: list, report_name: str) -> Path:
    wb = Workbook()

    # ── Summary Sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 22

    ws_sum["A1"] = report_name
    ws_sum["A1"].font = Font(bold=True, size=16, color="FF6B35")
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_sum["A2"].font = Font(color="888888", size=10)
    ws_sum["A3"] = f"Rows processed: {len(df)}"
    ws_sum["A3"].font = Font(color="888888", size=10)

    ws_sum["A5"] = "Column"
    ws_sum["B5"] = "Summary"
    for cell in [ws_sum["A5"], ws_sum["B5"]]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row = 6
    for col in df.columns:
        ws_sum.cell(row, 1, col).font = Font(bold=True, color="F0EEE8")
        if pd.api.types.is_numeric_dtype(df[col]):
            summary = f"sum={df[col].sum():.2f} | avg={df[col].mean():.2f} | max={df[col].max():.2f}"
        else:
            summary = f"{df[col].nunique()} unique values"
        ws_sum.cell(row, 2, summary).font = Font(color="888888", size=10)
        if row % 2 == 0:
            for c in [ws_sum.cell(row, 1), ws_sum.cell(row, 2)]:
                c.fill = ALT_FILL
        row += 1

    # ── Data Sheet ─────────────────────────────────────────────────────────────
    ws_data = wb.create_sheet("Data")
    ws_data.sheet_view.showGridLines = False

    for ci, col in enumerate(df.columns, 1):
        cell = ws_data.cell(1, ci, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws_data.column_dimensions[get_column_letter(ci)].width = max(len(str(col)) + 4, 14)

    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            c = ws_data.cell(ri, ci, val)
            c.font = Font(color="F0EEE8", size=10)
            if ri % 2 == 0:
                c.fill = ALT_FILL
            c.alignment = Alignment(horizontal="center")

    # ── Charts Sheet ───────────────────────────────────────────────────────────
    if chart_paths:
        ws_charts = wb.create_sheet("Charts")
        ws_charts.sheet_view.showGridLines = False
        ws_charts["A1"] = "Charts"
        ws_charts["A1"].font = TITLE_FONT

        anchor_row = 3
        for cp in chart_paths:
            img = XLImage(str(cp))
            img.width, img.height = 700, 350
            ws_charts.add_image(img, f"A{anchor_row}")
            anchor_row += 22

    path = OUTPUT_DIR / f"{report_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    log.info(f"Excel saved: {path.name}")
    return path


# ── PDF Report ─────────────────────────────────────────────────────────────────
class ReportPDF(FPDF):
    def __init__(self, report_name):
        super().__init__()
        self.report_name = report_name

    def header(self):
        self.set_fill_color(13, 13, 13)
        self.rect(0, 0, 210, 18, "F")
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(255, 107, 53)
        self.cell(0, 18, self.report_name, align="L", ln=True)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(136, 136, 136)
        self.cell(0, 10, f"Page {self.page_no()} | Generated {datetime.now().strftime('%Y-%m-%d')}", align="C")


def build_pdf(df: pd.DataFrame, chart_paths: list, report_name: str) -> Path:
    pdf = ReportPDF(report_name)
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_fill_color(13, 13, 13)
    pdf.rect(0, 0, 210, 297, "F")

    # Title block
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(255, 107, 53)
    pdf.ln(8)
    pdf.cell(0, 10, report_name, ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(136, 136, 136)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Rows: {len(df)}", ln=True)
    pdf.ln(6)

    # Summary stats
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(240, 238, 232)
    pdf.cell(0, 8, "Summary Statistics", ln=True)
    pdf.ln(2)

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if numeric_cols:
        col_w = 180 / (len(numeric_cols) + 1)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(255, 107, 53)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(col_w, 7, "Column", border=0, fill=True)
        for nc in numeric_cols:
            pdf.cell(col_w, 7, str(nc)[:14], border=0, fill=True)
        pdf.ln()

        stats = [("Sum", df[numeric_cols].sum()), ("Mean", df[numeric_cols].mean()), ("Max", df[numeric_cols].max())]
        for label, series in stats:
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(240, 238, 232)
            pdf.set_fill_color(30, 30, 30)
            pdf.cell(col_w, 6, label, border=0, fill=True)
            for nc in numeric_cols:
                pdf.cell(col_w, 6, f"{series[nc]:.2f}", border=0, fill=True)
            pdf.ln()

    pdf.ln(8)

    # Charts
    for cp in chart_paths:
        pdf.add_page()
        pdf.set_fill_color(13, 13, 13)
        pdf.rect(0, 0, 210, 297, "F")
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(240, 238, 232)
        pdf.ln(4)
        pdf.cell(0, 8, cp.stem.replace("_", " ").title(), ln=True)
        pdf.image(str(cp), x=10, w=190)

    path = OUTPUT_DIR / f"{report_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(str(path))
    log.info(f"PDF saved: {path.name}")
    return path


# ── Email ──────────────────────────────────────────────────────────────────────
def send_email(attachments: list):
    cfg = CONFIG.get("email", {})
    if not cfg.get("enabled", False):
        log.info("Email disabled in config, skipping.")
        return

    msg = MIMEMultipart()
    msg["From"]    = cfg["sender"]
    msg["To"]      = ", ".join(cfg["recipients"])
    msg["Subject"] = cfg.get("subject", f"Automated Report — {datetime.now().strftime('%Y-%m-%d')}")

    body = cfg.get("body", "Please find the automated report attached.")
    msg.attach(MIMEText(body, "plain"))

    for fpath in attachments:
        with open(fpath, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={Path(fpath).name}")
        msg.attach(part)

    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg.get("smtp_port", 587)) as server:
            server.starttls()
            server.login(cfg["sender"], cfg["password"])
            server.sendmail(cfg["sender"], cfg["recipients"], msg.as_string())
        log.info(f"Email sent to: {cfg['recipients']}")
    except Exception as e:
        log.error(f"Email failed: {e}")


# ── Main Job ───────────────────────────────────────────────────────────────────
def run_report():
    log.info("=== Report job started ===")
    try:
        df           = load_data(CONFIG["data_source"])
        chart_paths  = generate_charts(df)
        report_name  = CONFIG.get("report_name", "Automated Report")
        excel_path   = build_excel(df, chart_paths, report_name)
        pdf_path     = build_pdf(df, chart_paths, report_name)
        send_email([excel_path, pdf_path])
        log.info("=== Report job completed ===")
    except Exception as e:
        log.error(f"Report job failed: {e}")


# ── Scheduler ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    sched_cfg = CONFIG.get("schedule", {})
    mode      = sched_cfg.get("mode", "once")

    if mode == "once":
        run_report()

    elif mode == "daily":
        run_time = sched_cfg.get("at", "08:00")
        schedule.every().day.at(run_time).do(run_report)
        log.info(f"Scheduler running — daily at {run_time}. Press Ctrl+C to stop.")
        while True:
            schedule.run_pending()
            time.sleep(30)

    elif mode == "interval":
        hours = sched_cfg.get("hours", 6)
        schedule.every(hours).hours.do(run_report)
        log.info(f"Scheduler running — every {hours}h. Press Ctrl+C to stop.")
        while True:
            schedule.run_pending()
            time.sleep(30)

    elif mode == "weekly":
        day     = sched_cfg.get("day", "monday")
        at_time = sched_cfg.get("at", "08:00")
        getattr(schedule.every(), day).at(at_time).do(run_report)
        log.info(f"Scheduler running — every {day} at {at_time}.")
        while True:
            schedule.run_pending()
            time.sleep(30)
